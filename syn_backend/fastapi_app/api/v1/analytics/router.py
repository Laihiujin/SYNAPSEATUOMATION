import csv
import io
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Query, HTTPException, Body
from pydantic import BaseModel, Field
from fastapi.responses import StreamingResponse
from fastapi_app.core.config import settings
from myUtils.analytics_db import (
    ensure_analytics_schema,
    get_analytics_summary,
    get_analytics_videos,
    get_chart_data,
    insert_video_analytics,
    update_video_analytics,
    upsert_video_analytics_by_key,
)
from fastapi_app.services.social_media_copilot import get_social_media_copilot_client
from fastapi_app.services.analytics_extract import extract_video_info
from datetime import datetime
from typing import Any, Dict, List
import sys
from myUtils.cookie_manager import cookie_manager

router = APIRouter(prefix="/analytics", tags=["数据分析"])

DB_PATH: Path = Path(settings.BASE_DIR) / "db" / "database.db"

# Ensure DB schema exists on import
ensure_analytics_schema(DB_PATH)


class CollectTask(BaseModel):
    platform: str = Field(..., description="平台：douyin/xiaohongshu/kuaishou")
    work_id: str = Field(..., description="作品 ID（aweme_id / note_id / photoId）")


class CollectPayload(BaseModel):
    mode: Optional[str] = Field(None, description="采集模式 accounts/works，默认按账号采集")
    platform: Optional[str] = Field(None, description="平台，可选：douyin/xiaohongshu/kuaishou/channels/all")
    account_ids: Optional[List[str]] = Field(None, description="账号ID列表，空为所有有效账号")
    work_ids: Optional[List[str]] = Field(None, description="作品 ID 列表，若提供 platform 则全部使用该平台")
    tasks: Optional[List[CollectTask]] = Field(None, description="自定义任务列表，含平台与作品 ID")



def _ensure_douyin_api_on_path() -> None:
    douyin_root = Path(settings.BASE_DIR) / "douyin_tiktok_api"
    if douyin_root.exists():
        sys.path.insert(0, str(douyin_root))




def _extract_bilibili_uid(cookie_data: Any) -> Optional[str]:
    if not isinstance(cookie_data, dict):
        return None
    cookies = cookie_data.get("cookies")
    if not cookies and isinstance(cookie_data.get("cookie_info"), dict):
        cookies = cookie_data.get("cookie_info", {}).get("cookies")
    if not isinstance(cookies, list):
        return None
    for cookie in cookies:
        if isinstance(cookie, dict) and cookie.get("name") == "DedeUserID":
            value = cookie.get("value")
            if value:
                return str(value)
    return None

def _parse_bilibili_vlist(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    data = payload.get("data") if isinstance(payload, dict) else None
    if isinstance(data, dict) and "list" in data:
        data = data.get("list")
    if isinstance(data, dict):
        vlist = data.get("vlist")
        if isinstance(vlist, list):
            return vlist
    return []


def _to_date_str(ts: Any) -> str:
    try:
        ts_int = int(ts)
        if ts_int > 1e12:
            ts_int = ts_int / 1000
        return datetime.utcfromtimestamp(ts_int).date().isoformat()
    except Exception:
        return datetime.utcnow().date().isoformat()


async def _collect_bilibili_accounts(account_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    _ensure_douyin_api_on_path()
    try:
        from crawlers.bilibili.web.web_crawler import BilibiliWebCrawler  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Bilibili crawler import failed: {exc}")

    accounts = [
        acc for acc in cookie_manager.list_flat_accounts()
        if acc.get("platform") == "bilibili"
    ]
    if account_ids:
        accounts = [acc for acc in accounts if acc.get("account_id") in set(account_ids)]

    crawler = BilibiliWebCrawler()
    success = 0
    failed: List[Dict[str, Any]] = []

    for account in accounts:
        account_id = account.get("account_id")
        cookie_data = cookie_manager._read_cookie_file(account.get("cookie_file") or "")
        uid = (
            account.get("user_id")
            or cookie_manager._extract_user_id_from_cookie("bilibili", cookie_data)
            or _extract_bilibili_uid(cookie_data)
        )

        if not uid:
            failed.append({"account_id": account_id, "error": "missing uid"})
            continue

        try:
            pn = 1
            max_pages = 5
            while pn <= max_pages:
                payload = await crawler.fetch_user_post_videos(uid=str(uid), pn=pn)
                vlist = _parse_bilibili_vlist(payload)
                if not vlist:
                    break
                for item in vlist:
                    video_id = item.get("bvid") or item.get("aid")
                    if not video_id:
                        continue
                    bvid = item.get("bvid")
                    aid = item.get("aid")
                    video_url = None
                    if bvid:
                        video_url = f"https://www.bilibili.com/video/{bvid}"
                    elif aid:
                        video_url = f"https://www.bilibili.com/video/av{aid}"

                    record = {
                        "account_id": account_id,
                        "platform": "bilibili",
                        "video_id": str(video_id),
                        "video_url": video_url,
                        "title": item.get("title") or "",
                        "thumbnail": item.get("pic"),
                        "publish_date": _to_date_str(item.get("created")),
                        "play_count": item.get("play") or 0,
                        "like_count": item.get("like") or 0,
                        "comment_count": item.get("comment") or 0,
                        "collect_count": item.get("favorites") or 0,
                        "share_count": item.get("share") or 0,
                        "raw_data": item,
                        "collected_at": datetime.utcnow().isoformat(),
                    }
                    upsert_video_analytics_by_key(DB_PATH, platform="bilibili", video_id=str(video_id), data=record)
                pn += 1
            success += 1
        except Exception as exc:  # noqa: BLE001
            failed.append({"account_id": account_id, "error": str(exc)})

    return {"success": success, "failed": len(failed), "errors": failed}

@router.get("/", summary="获取分析数据")
async def get_analytics(
    startDate: Optional[str] = Query(None, description="???? YYYY-MM-DD"),
    endDate: Optional[str] = Query(None, description="???? YYYY-MM-DD"),
    platform: Optional[str] = Query(None, description="????"),
    limit: int = Query(100, ge=1, le=10000, description="?????")
):
    """获取汇总、视频列表和图表数据"""
    try:
        summary = get_analytics_summary(DB_PATH, startDate, endDate)
        videos = get_analytics_videos(DB_PATH, startDate, endDate, limit, platform)
        chart_data = get_chart_data(DB_PATH, startDate, endDate)
        return {
            "code": 200,
            "summary": summary,
            "videos": videos,
            "chartData": chart_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export", summary="导出分析数据")
async def export_analytics(
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    format: str = Query("csv", pattern="^(csv|excel)$", description="???? csv ? excel")
):
    """导出分析数据为 CSV（默认）或 Excel（若依赖存在）"""
    try:
        videos = get_analytics_videos(DB_PATH, startDate, endDate, limit=10000, platform=platform)
        if format == "excel":
            try:
                import openpyxl  # type: ignore
                from openpyxl.styles import Font, Alignment  # type: ignore
            except ImportError:
                format = "csv"

        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'ID', '视频ID', '标题', '平台', '视频链接',
                '发布日期', '播放量', '点赞量', '评论量', '收藏量', '最后更新'
            ])
            for video in videos:
                writer.writerow([
                    video.get('id'),
                    video.get('videoId'),
                    video.get('title'),
                    video.get('platform'),
                    video.get('videoUrl', ''),
                    video.get('publishDate'),
                    video.get('playCount', 0),
                    video.get('likeCount', 0),
                    video.get('commentCount', 0),
                    video.get('collectCount', 0),
                    video.get('lastUpdated')
                ])
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": "attachment; filename=analytics_export.csv",
                    "Content-Type": "text/csv; charset=utf-8-sig"
                }
            )
        else:
            # Excel 导出
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "视频数据"
            headers = [
                'ID', '视频ID', '标题', '平台', '视频链接',
                '发布日期', '播放量', '点赞量', '评论量', '收藏量', '最后更新'
            ]
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
            for row_idx, video in enumerate(videos, 2):
                ws.cell(row=row_idx, column=1, value=video.get('id'))
                ws.cell(row=row_idx, column=2, value=video.get('videoId'))
                ws.cell(row=row_idx, column=3, value=video.get('title'))
                ws.cell(row=row_idx, column=4, value=video.get('platform'))
                ws.cell(row=row_idx, column=5, value=video.get('videoUrl', ''))
                ws.cell(row=row_idx, column=6, value=video.get('publishDate'))
                ws.cell(row=row_idx, column=7, value=video.get('playCount', 0))
                ws.cell(row=row_idx, column=8, value=video.get('likeCount', 0))
                ws.cell(row=row_idx, column=9, value=video.get('commentCount', 0))
                ws.cell(row=row_idx, column=10, value=video.get('collectCount', 0))
                ws.cell(row=row_idx, column=11, value=video.get('lastUpdated'))
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=analytics_export.xlsx"}
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/video", summary="新增视频分析记录")
async def add_video_analytics(payload: dict):
    """新增视频分析记录"""
    try:
        video_id = insert_video_analytics(DB_PATH, payload)
        return {"code": 200, "msg": "Success", "videoId": video_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/video/{video_id}", summary="更新视频分析数据")
async def update_video(video_id: int, payload: dict):
    """更新视频分析数据"""
    try:
        update_video_analytics(DB_PATH, video_id, payload)
        return {"code": 200, "msg": "Updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/collect/{platform}", summary="Platform-specific analytics collection (Bilibili ready)")
async def collect_platform(platform: str, payload: Optional[CollectPayload] = Body(default=None)):
    platform = (platform or "").lower()
    account_ids = payload.account_ids if payload else None

    if platform == "bilibili":
        result = await _collect_bilibili_accounts(account_ids=account_ids)
        return {"success": True, "data": result, "message": "Bilibili collection completed"}

    raise HTTPException(status_code=501, detail="Not implemented")


@router.post("/collect", summary="?? copilot ??????? analytics")
async def collect_analytics(payload: CollectPayload):
    """
    - ????????????????????? Cookie?
    - ?????? tasks ? platform + work_ids????? ID ??
    """
    collect_mode = (payload.mode or "").lower()
    collect_by_account = collect_mode in {"account", "accounts", "by_account", "all_accounts"}
    collect_by_account = collect_by_account or payload.account_ids is not None
    collect_by_account = collect_by_account or (not payload.tasks and not payload.work_ids)

    if collect_by_account and (payload.platform or "").lower() == "bilibili":
        result = await _collect_bilibili_accounts(account_ids=payload.account_ids)
        return {
            "success": True,
            "data": result,
            "message": "Bilibili collection completed",
        }

    if collect_by_account:
        try:
            from myUtils.video_collector import collector
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"Collector init failed: {exc}")

        platform_filter = None
        if payload.platform and payload.platform.lower() != "all":
            platform_filter = payload.platform.lower()

        results = await collector.collect_all_accounts(
            account_ids=payload.account_ids,
            platform_filter=platform_filter,
        )
        return {
            "success": True,
            "data": results,
            "message": "Account collection completed",
        }

    tasks: List[CollectTask] = []

    if payload.tasks:
        tasks = payload.tasks
    elif payload.platform and payload.work_ids:
        tasks = [CollectTask(platform=payload.platform, work_id=w) for w in payload.work_ids]
    else:
        raise HTTPException(status_code=400, detail="??? tasks ? platform + work_ids")

    client = get_social_media_copilot_client()
    success_count = 0
    failures: List[Dict[str, Any]] = []

    for task in tasks:
        try:
            result = await client.fetch_work(task.platform, task.work_id)
            if not result.get("success"):
                failures.append(
                    {"platform": task.platform, "work_id": task.work_id, "error": result.get("error")}
                )
                continue

            record = extract_video_info(task.platform, result.get("data") or {}, task.work_id)
            upsert_video_analytics_by_key(DB_PATH, platform=record.get("platform") or task.platform, video_id=record.get("video_id") or task.work_id, data=record)
            success_count += 1
        except Exception as exc:  # noqa: BLE001
            failures.append({"platform": task.platform, "work_id": task.work_id, "error": str(exc)})

    return {
        "success": len(failures) == 0,
        "data": {"success": success_count, "failed": len(failures)},
        "errors": failures,
    }
